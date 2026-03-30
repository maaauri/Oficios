"""Herramienta independiente para corregir área responsable y plazo de oficios ya procesados."""

from __future__ import annotations

import argparse
import json
import logging
import sys
import tkinter as tk
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from tkinter import messagebox
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment


# ---------------------------------------------------------------------------
# Configuración mínima
# ---------------------------------------------------------------------------

AREAS_VALIDAS = ["PMGD", "Conexiones", "Lectura", "Servicio al Cliente", "Cobranza", "Pérdidas"]


@dataclass
class Gerente:
    nombre: str
    email: str = ""


def load_revaluar_config(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        raw = json.load(f)
    return {
        "excel_path": Path(raw["excel_path"]),
        "corrections_path": Path(raw.get("corrections_path", "corrections.json")),
        "log_path": Path(raw.get("log_path", "oficios_service.log")),
        "gerentes": {
            area: Gerente(nombre=item.get("nombre", ""), email=item.get("email", ""))
            for area, item in raw.get("gerentes", {}).items()
        },
    }


# ---------------------------------------------------------------------------
# Funciones auxiliares
# ---------------------------------------------------------------------------

def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def parse_date_yyyy_mm_dd(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def load_corrections(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_corrections(path: Path, corrections: List[Dict[str, Any]]) -> None:
    ensure_parent(path)
    with path.open("w", encoding="utf-8") as f:
        json.dump(corrections, f, ensure_ascii=False, indent=2)


def update_excel_row(
    excel_path: Path, nro: str, categoria: str,
    updates: Dict[str, Any], gerentes: Dict[str, Gerente],
) -> bool:
    wb = load_workbook(excel_path)
    ws = wb.active
    for row_idx in range(2, ws.max_row + 1):
        cell_nro = str(ws.cell(row=row_idx, column=1).value or "").strip()
        cell_cat = str(ws.cell(row=row_idx, column=2).value or "").strip()
        if cell_nro == nro and cell_cat == categoria:
            if "gerencia_responsable" in updates:
                new_gerencia = updates["gerencia_responsable"]
                ws.cell(row=row_idx, column=6, value=new_gerencia)
                gerente = gerentes.get(new_gerencia, Gerente(nombre=""))
                ws.cell(row=row_idx, column=7, value=gerente.nombre)
            if "plazo_respuesta" in updates:
                new_plazo = parse_date_yyyy_mm_dd(updates["plazo_respuesta"])
                if new_plazo:
                    cell = ws.cell(row=row_idx, column=9, value=new_plazo)
                    cell.number_format = "DD-MM-YYYY"
                    cell.alignment = Alignment(horizontal="center")
            wb.save(excel_path)
            return True
    return False


# ---------------------------------------------------------------------------
# Interfaz gráfica
# ---------------------------------------------------------------------------

def show_revaluar_gui(
    excel_path: Path, corrections_path: Path, gerentes: Dict[str, Gerente],
) -> None:
    if not excel_path.exists():
        logging.error("No existe el Excel: %s", excel_path)
        return

    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nro = str(row[0] or "").strip()
        if not nro:
            continue
        cat = str(row[1] or "")
        concepto = str(row[3] or "")
        gerencia = str(row[5] or "")
        plazo = row[8]
        if hasattr(plazo, "date"):
            plazo = plazo.date()
        plazo_str = plazo.strftime("%d-%m-%Y") if isinstance(plazo, date) else ""
        rows_data.append({
            "nro": nro, "categoria": cat, "concepto": concepto,
            "gerencia": gerencia, "plazo_str": plazo_str,
        })
    wb.close()

    if not rows_data:
        messagebox.showinfo("Revaloración", "No hay oficios en el Excel.")
        return

    root = tk.Tk()
    root.title("Revaloración de oficios")
    root.geometry("800x500")
    root.resizable(True, True)

    # --- Lista de oficios ---
    frame_list = tk.Frame(root)
    frame_list.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

    tk.Label(frame_list, text="Seleccione un oficio para corregir:", font=("Arial", 10, "bold")).pack(anchor=tk.W)

    listbox = tk.Listbox(frame_list, font=("Consolas", 9), selectmode=tk.SINGLE)
    scrollbar = tk.Scrollbar(frame_list, orient=tk.VERTICAL, command=listbox.yview)
    listbox.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    listbox.pack(fill=tk.BOTH, expand=True)

    for rd in rows_data:
        listbox.insert(tk.END, f"Nro {rd['nro']}  |  {rd['categoria']}  |  {rd['gerencia']}  |  {rd['plazo_str']}  |  {rd['concepto'][:60]}")

    # --- Formulario de corrección ---
    frame_form = tk.Frame(root)
    frame_form.pack(fill=tk.X, padx=10, pady=5)

    tk.Label(frame_form, text="Nueva área responsable:").grid(row=0, column=0, sticky=tk.W, pady=2)
    area_var = tk.StringVar(root)
    area_var.set("")
    area_options = ["(sin cambio)"] + AREAS_VALIDAS
    area_menu = tk.OptionMenu(frame_form, area_var, *area_options)
    area_menu.config(width=25)
    area_menu.grid(row=0, column=1, sticky=tk.W, padx=5)

    tk.Label(frame_form, text="Nuevo plazo (DD-MM-YYYY):").grid(row=1, column=0, sticky=tk.W, pady=2)
    plazo_entry = tk.Entry(frame_form, width=20)
    plazo_entry.grid(row=1, column=1, sticky=tk.W, padx=5)

    status_label = tk.Label(root, text="", fg="green", font=("Arial", 9))
    status_label.pack(pady=2)

    def on_save():
        sel = listbox.curselection()
        if not sel:
            messagebox.showwarning("Revaloración", "Seleccione un oficio primero.")
            return
        idx = sel[0]
        rd = rows_data[idx]
        corrections = load_corrections(corrections_path)
        updates: Dict[str, Any] = {}
        new_area = area_var.get()
        new_plazo_raw = plazo_entry.get().strip()

        if new_area and new_area != "(sin cambio)":
            corrections.append({
                "nro": rd["nro"],
                "campo": "gerencia_responsable",
                "valor_anterior": rd["gerencia"],
                "valor_nuevo": new_area,
                "concepto": rd["concepto"],
            })
            updates["gerencia_responsable"] = new_area

        if new_plazo_raw:
            try:
                parsed_plazo = datetime.strptime(new_plazo_raw, "%d-%m-%Y").date()
                corrections.append({
                    "nro": rd["nro"],
                    "campo": "plazo_respuesta",
                    "valor_anterior": rd["plazo_str"],
                    "valor_nuevo": parsed_plazo.isoformat(),
                    "concepto": rd["concepto"],
                })
                updates["plazo_respuesta"] = parsed_plazo.isoformat()
            except ValueError:
                messagebox.showerror("Error", "Formato de fecha inválido. Use DD-MM-YYYY.")
                return

        if not updates:
            messagebox.showinfo("Revaloración", "No se indicó ninguna corrección.")
            return

        update_excel_row(excel_path, rd["nro"], rd["categoria"], updates, gerentes)
        save_corrections(corrections_path, corrections)

        # Actualizar listbox
        new_gerencia = updates.get("gerencia_responsable", rd["gerencia"])
        new_plazo_str = rd["plazo_str"]
        if "plazo_respuesta" in updates:
            p = parse_date_yyyy_mm_dd(updates["plazo_respuesta"])
            new_plazo_str = p.strftime("%d-%m-%Y") if p else rd["plazo_str"]
        rd["gerencia"] = new_gerencia
        rd["plazo_str"] = new_plazo_str
        listbox.delete(idx)
        listbox.insert(idx, f"Nro {rd['nro']}  |  {rd['categoria']}  |  {rd['gerencia']}  |  {rd['plazo_str']}  |  {rd['concepto'][:60]}")

        status_label.config(text=f"Oficio Nro {rd['nro']} corregido correctamente.")
        logging.info("Corrección aplicada: Nro %s — %s", rd["nro"], updates)

        area_var.set("(sin cambio)")
        plazo_entry.delete(0, tk.END)

    tk.Button(root, text="Guardar corrección", command=on_save, bg="#1F4E78", fg="white", font=("Arial", 10, "bold")).pack(pady=8)

    root.mainloop()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Corrige área responsable y plazo de oficios ya procesados.",
    )
    parser.add_argument("--config", default="config.json", help="Ruta al archivo de configuración JSON.")
    args = parser.parse_args()

    cfg = load_revaluar_config(Path(args.config))

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(cfg["log_path"], encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )

    show_revaluar_gui(cfg["excel_path"], cfg["corrections_path"], cfg["gerentes"])


if __name__ == "__main__":
    main()
