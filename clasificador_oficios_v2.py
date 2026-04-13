#!/usr/bin/env python3
"""
Clasificador de Oficios CGE — v2 (Agente con Aprendizaje)
==========================================================
Escanea el directorio actual buscando PDFs de oficios (Ord*, OC*, RE*),
los clasifica con IA y genera/actualiza un Excel con planner.

NUEVO en v2:
  - Feedback loop: después de cada clasificación puedes confirmar o corregir
  - Historial persistente: cada decisión se guarda en historial_oficios.json
  - Few-shot dinámico: el prompt incluye los ejemplos más relevantes del historial
  - Reglas aprendidas: se generan automáticamente cada N oficios procesados
  - Métricas de accuracy por área

Requisitos:
    pip install anthropic openpyxl

Uso:
    1. Coloca este script en la carpeta donde están los oficios PDF
    2. Configura tu API key de Anthropic (ver abajo)
    3. Ejecuta: python clasificador_oficios.py

Comandos especiales:
    python clasificador_oficios.py --stats       Ver métricas de accuracy
    python clasificador_oficios.py --regenerar   Regenerar reglas aprendidas manualmente
"""

import os
import sys
import json
import base64
import glob
import argparse
from datetime import datetime, date
from collections import Counter

try:
    import anthropic
except ImportError:
    print("ERROR: Instala la librería anthropic:")
    print("  pip install anthropic")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: Instala openpyxl:")
    print("  pip install openpyxl")
    sys.exit(1)

# ============================================================
# CONFIGURACIÓN
# ============================================================

API_KEY = ""  # Pega tu API key aquí, o usa variable de entorno ANTHROPIC_API_KEY

EXCEL_FILE = "CGE_Oficios_Planner.xlsx"
SHEET_OFICIOS = "Oficios"
SHEET_PLANNER = "Planner"

# Archivos del sistema de aprendizaje
HISTORIAL_FILE = "historial_oficios.json"
REGLAS_FILE = "reglas_clasificacion.json"

# Cada cuántos oficios se regeneran las reglas automáticamente
REGLAS_CADA_N = 20

# Cuántos ejemplos few-shot incluir en el prompt
MAX_FEWSHOT = 10

AREAS_REPRESENTANTES = {
    "PMGD": ["Nirvia", "Edgar"],
    "Conexiones": ["Claudia Cea", "Maricela Villanueva", "Carolina González"],
    "Lectura y Facturación": ["Claudio Osorio", "Roberto Domínguez", "Carlos Felipe Mora"],
    "Reclamos": ["Sandra Cortés", "Jessica Montecinos", "Ruby Mallea"],
    "Cobranza": ["Karen Carvajal"],
}

AREAS_VALIDAS = list(AREAS_REPRESENTANTES.keys())

HEADER_BG = "001689"
HEADER_FONT_COLOR = "FFFFFF"


# ============================================================
# SISTEMA DE APRENDIZAJE
# ============================================================

def load_historial():
    """Carga el historial de clasificaciones confirmadas/corregidas."""
    if not os.path.exists(HISTORIAL_FILE):
        return []
    try:
        with open(HISTORIAL_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"  ⚠️ Error al leer historial: {e}")
        return []


def save_historial(historial):
    """Guarda el historial de clasificaciones."""
    with open(HISTORIAL_FILE, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)


def add_to_historial(historial, oficio, area_propuesta, area_final, representante):
    """Agrega una entrada al historial con metadata de corrección."""
    entry = {
        "archivo": oficio.get("archivo", ""),
        "nombre_oficio": oficio.get("nombre_oficio", ""),
        "resumen": oficio.get("resumen", ""),
        "remitente": oficio.get("remitente", ""),
        "area_propuesta": area_propuesta,
        "area_final": area_final,
        "fue_corregido": area_propuesta != area_final,
        "representante": representante,
        "fecha_procesado": datetime.now().isoformat(),
    }
    historial.append(entry)
    save_historial(historial)
    return entry


def build_fewshot_examples(historial, max_examples=MAX_FEWSHOT):
    """
    Construye ejemplos few-shot del historial.
    Prioriza:
      1. Correcciones (donde el modelo se equivocó) — son las más valiosas
      2. Los más recientes
    """
    if not historial:
        return ""

    # Separar correcciones y confirmaciones
    correcciones = [h for h in historial if h.get("fue_corregido")]
    confirmaciones = [h for h in historial if not h.get("fue_corregido")]

    # Tomar más correcciones que confirmaciones (ratio ~60/40)
    n_correcciones = min(len(correcciones), max(1, int(max_examples * 0.6)))
    n_confirmaciones = min(len(confirmaciones), max_examples - n_correcciones)

    # Más recientes primero
    selected = correcciones[-n_correcciones:] + confirmaciones[-n_confirmaciones:]

    if not selected:
        return ""

    lines = ["", "EJEMPLOS DE CLASIFICACIONES ANTERIORES CONFIRMADAS POR EL USUARIO:"]
    lines.append("(Usa estos como referencia para clasificar correctamente)")
    lines.append("")

    for i, ex in enumerate(selected, 1):
        tag = "⚠️ CORRECCIÓN" if ex.get("fue_corregido") else "✓ Confirmado"
        lines.append(f"  Ejemplo {i} [{tag}]:")
        lines.append(f"    Oficio: {ex.get('nombre_oficio', '?')}")
        lines.append(f"    Resumen: {ex.get('resumen', '?')}")
        if ex.get("fue_corregido"):
            lines.append(f"    Área propuesta (INCORRECTA): {ex.get('area_propuesta', '?')}")
            lines.append(f"    Área correcta: {ex.get('area_final', '?')}")
        else:
            lines.append(f"    Área: {ex.get('area_final', '?')}")
        lines.append("")

    return "\n".join(lines)


def load_reglas():
    """Carga las reglas de clasificación aprendidas."""
    if not os.path.exists(REGLAS_FILE):
        return None
    try:
        with open(REGLAS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data.get("reglas_texto", "")
    except Exception:
        return None


def save_reglas(reglas_texto, stats):
    """Guarda las reglas de clasificación generadas."""
    data = {
        "generado": datetime.now().isoformat(),
        "basado_en_n_oficios": stats.get("total", 0),
        "accuracy_global": stats.get("accuracy", 0),
        "reglas_texto": reglas_texto,
    }
    with open(REGLAS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def generate_reglas(client, historial):
    """
    Usa Claude para analizar el historial completo y generar reglas
    de clasificación optimizadas.
    """
    if len(historial) < 5:
        print("  ℹ️ Se necesitan al menos 5 oficios procesados para generar reglas.")
        return None

    # Preparar resumen del historial para el análisis
    resumen_historial = []
    for h in historial:
        entry = {
            "oficio": h.get("nombre_oficio", ""),
            "resumen": h.get("resumen", ""),
            "area_final": h.get("area_final", ""),
            "fue_corregido": h.get("fue_corregido", False),
        }
        if h.get("fue_corregido"):
            entry["area_propuesta_incorrecta"] = h.get("area_propuesta", "")
        resumen_historial.append(entry)

    prompt = f"""Eres un analista de documentos oficiales de la SEC (Superintendencia de Electricidad y Combustibles) de Chile,
dirigidos a CGE (Compañía General de Electricidad).

Analiza el siguiente historial de clasificaciones de oficios y genera REGLAS DE CLASIFICACIÓN optimizadas.

Las áreas posibles son: {', '.join(AREAS_VALIDAS)}

HISTORIAL:
{json.dumps(resumen_historial, ensure_ascii=False, indent=2)}

Genera reglas de clasificación que:
1. Identifiquen patrones claros de keywords, temas y remitentes por cada área
2. Presten especial atención a las CORRECCIONES (donde el modelo se equivocó)
3. Sean específicas y accionables (no genéricas)
4. Incluyan reglas negativas cuando sea útil ("NO clasificar como X si...")

Responde SOLO con las reglas en texto plano, sin JSON ni markdown. Formato:

ÁREA: [nombre]
- Regla 1
- Regla 2
...

Incluye una sección final de REGLAS GENERALES si hay patrones transversales."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
        )
        reglas = "".join(block.text for block in response.content if hasattr(block, "text"))

        stats = compute_stats(historial)
        save_reglas(reglas, stats)
        return reglas
    except Exception as e:
        print(f"  ⚠️ Error al generar reglas: {e}")
        return None


def compute_stats(historial):
    """Calcula métricas de accuracy del agente."""
    if not historial:
        return {"total": 0, "accuracy": 0, "por_area": {}}

    total = len(historial)
    correctos = sum(1 for h in historial if not h.get("fue_corregido"))
    accuracy = correctos / total if total > 0 else 0

    # Accuracy por área
    por_area = {}
    for area in AREAS_VALIDAS:
        area_entries = [h for h in historial if h.get("area_final") == area]
        if area_entries:
            area_correctos = sum(1 for h in area_entries if not h.get("fue_corregido"))
            por_area[area] = {
                "total": len(area_entries),
                "correctos": area_correctos,
                "accuracy": area_correctos / len(area_entries),
            }

    # Errores más frecuentes
    errores = [
        f"{h.get('area_propuesta')} → {h.get('area_final')}"
        for h in historial if h.get("fue_corregido")
    ]
    errores_frecuentes = Counter(errores).most_common(5)

    return {
        "total": total,
        "correctos": correctos,
        "accuracy": accuracy,
        "por_area": por_area,
        "errores_frecuentes": errores_frecuentes,
    }


def print_stats(historial):
    """Muestra las métricas de accuracy en consola."""
    stats = compute_stats(historial)

    print()
    print("╔══════════════════════════════════════════════════════════╗")
    print("║           MÉTRICAS DEL AGENTE CLASIFICADOR              ║")
    print("╚══════════════════════════════════════════════════════════╝")
    print()
    print(f"  Total oficios procesados:  {stats['total']}")
    print(f"  Clasificados correctamente: {stats['correctos']}")
    print(f"  Accuracy global:           {stats['accuracy']:.1%}")
    print()

    if stats["por_area"]:
        print("  Accuracy por área:")
        print("  ─────────────────────────────────────────")
        for area, data in stats["por_area"].items():
            bar = "█" * int(data["accuracy"] * 20) + "░" * (20 - int(data["accuracy"] * 20))
            print(f"    {area:<25} {bar} {data['accuracy']:.0%}  ({data['correctos']}/{data['total']})")
        print()

    if stats.get("errores_frecuentes"):
        print("  Errores más frecuentes (propuesta → corrección):")
        print("  ─────────────────────────────────────────")
        for error, count in stats["errores_frecuentes"]:
            print(f"    {error:<40} x{count}")
        print()

    # Info de reglas
    if os.path.exists(REGLAS_FILE):
        try:
            with open(REGLAS_FILE, "r", encoding="utf-8") as f:
                reglas_data = json.load(f)
            print(f"  Reglas generadas: {reglas_data.get('generado', '?')}")
            print(f"  Basadas en: {reglas_data.get('basado_en_n_oficios', '?')} oficios")
            print(f"  Accuracy al generar: {reglas_data.get('accuracy_global', 0):.1%}")
        except Exception:
            pass
    else:
        print("  Reglas: No generadas aún (se generan automáticamente cada "
              f"{REGLAS_CADA_N} oficios)")
    print()


# ============================================================
# FUNCIONES BASE (del v1, mejoradas)
# ============================================================

def get_api_key():
    key = API_KEY or os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        print("=" * 60)
        print("  API KEY NO CONFIGURADA")
        print("=" * 60)
        print()
        print("Necesitas una API key de Anthropic para usar este script.")
        print()
        print("Opciones:")
        print("  1. Edita este archivo y pega tu key en API_KEY = \"...\"")
        print("  2. Configura la variable de entorno:")
        print('     export ANTHROPIC_API_KEY="sk-ant-..."')
        print()
        print("Obtén tu key en: https://console.anthropic.com/")
        sys.exit(1)
    return key


def find_oficios_pdfs():
    """Busca PDFs que empiecen con Ord, OC o RE en el directorio actual."""
    patterns = ["Ord*.pdf", "OC*.pdf", "RE*.pdf",
                "Ord*.PDF", "OC*.PDF", "RE*.PDF",
                "ORD*.pdf", "ORD*.PDF"]
    found = set()
    for p in patterns:
        found.update(glob.glob(p))
    return sorted(found)


def load_existing_oficios():
    """Carga los oficios ya procesados del Excel existente."""
    existing = set()
    if not os.path.exists(EXCEL_FILE):
        return existing
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        if SHEET_OFICIOS in wb.sheetnames:
            ws = wb[SHEET_OFICIOS]
            for row in ws.iter_rows(min_row=2, values_only=True):
                archivo = row[0] if row[0] else ""
                existing.add(str(archivo))
        wb.close()
    except Exception as e:
        print(f"  ⚠️ Advertencia al leer Excel existente: {e}")
    return existing


def classify_pdf(client, filepath, historial):
    """
    Envía el PDF a Claude para clasificación.
    Incluye few-shot dinámico del historial y reglas aprendidas.
    """
    with open(filepath, "rb") as f:
        pdf_bytes = f.read()
    b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

    # Construir las secciones dinámicas del prompt
    fewshot_section = build_fewshot_examples(historial)
    reglas_section = ""
    reglas_aprendidas = load_reglas()
    if reglas_aprendidas:
        reglas_section = f"""

REGLAS DE CLASIFICACIÓN APRENDIDAS (basadas en decisiones anteriores del usuario):
{reglas_aprendidas}

Estas reglas tienen prioridad sobre las reglas genéricas cuando haya conflicto.
"""

    prompt = f"""Analiza este oficio/documento oficial chileno y extrae la siguiente información.
Responde SOLO con un objeto JSON válido, sin markdown ni backticks ni texto adicional.

Campos requeridos:
- "nombre_oficio": string — nombre, número o identificador del oficio (ej: "Oficio N°123", "Res. Ex. N°456")
- "fecha_oficio": string en formato YYYY-MM-DD — la fecha del documento
- "fecha_plazo": string en formato YYYY-MM-DD o null — fecha límite de respuesta/cumplimiento si existe
- "area": string — debe ser EXACTAMENTE una de: {', '.join(f'"{a}"' for a in AREAS_VALIDAS)}
  
  REGLAS BASE DE CLASIFICACIÓN:
  * PMGD: SOLO si el documento menciona explícitamente "PMGD", "Pequeño Medio de Generación Distribuida" o "PMG". Si no aparece ninguna de estas siglas/términos, NO clasificar como PMGD aunque trate de generación o energía.
  * Conexiones: solicitudes de conexión, empalmes, factibilidad técnica, alimentadores, obras eléctricas
  * Lectura y Facturación: lectura de medidores, facturación, consumos, tarifas, medidores
  * Reclamos: reclamos de clientes, quejas, solicitudes de servicio al cliente, SEC reclamos
  * Cobranza: cobros, deudas, cortes, reposiciones, morosidad, convenios de pago
{reglas_section}
- "confianza": number entre 0.0 y 1.0 — qué tan seguro estás de la clasificación del área
- "resumen": string — breve resumen del oficio en máximo 2 líneas
- "remitente": string — quién envía el oficio
- "keywords": list[string] — 3-5 palabras clave del documento que justifican la clasificación
{fewshot_section}

Responde SOLO el JSON."""

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1200,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )

    text = "".join(block.text for block in response.content if hasattr(block, "text"))
    clean = text.replace("```json", "").replace("```", "").strip()
    return json.loads(clean)


def format_date_cl(d):
    """Convierte YYYY-MM-DD a DD/MM/YYYY."""
    if not d:
        return ""
    try:
        dt = datetime.strptime(d, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return str(d)


def dias_restantes(fecha_plazo):
    if not fecha_plazo:
        return None
    try:
        plazo = datetime.strptime(fecha_plazo, "%Y-%m-%d").date()
        return (plazo - date.today()).days
    except Exception:
        return None


def estado_oficio(fecha_plazo):
    dias = dias_restantes(fecha_plazo)
    if dias is None:
        return "Sin plazo"
    if dias < 0:
        return f"Vencido ({abs(dias)}d)"
    if dias <= 3:
        return f"Urgente ({dias}d)"
    return f"Pendiente ({dias}d)"


def style_header(ws, num_cols):
    """Aplica estilo al encabezado: fondo #001689, letras blancas, negrita."""
    fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    font = Font(bold=True, color=HEADER_FONT_COLOR, name="Arial", size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="000000"))
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border


def build_excel(all_oficios):
    """Genera el Excel con hojas Oficios y Planner."""
    wb = Workbook()

    # --- Hoja Oficios ---
    ws = wb.active
    ws.title = SHEET_OFICIOS

    headers = ["Archivo", "Nombre Oficio", "Remitente", "Fecha Oficio",
               "Fecha Plazo", "Estado", "Área", "Representante", "Resumen"]
    ws.append(headers)
    style_header(ws, len(headers))

    for o in all_oficios:
        ws.append([
            o.get("archivo", ""),
            o.get("nombre_oficio", ""),
            o.get("remitente", ""),
            format_date_cl(o.get("fecha_oficio")),
            format_date_cl(o.get("fecha_plazo")),
            estado_oficio(o.get("fecha_plazo")),
            o.get("area", ""),
            o.get("representante", ""),
            o.get("resumen", ""),
        ])

    col_widths = [30, 28, 22, 14, 14, 18, 24, 26, 55]
    letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    for i, letter in enumerate(letters):
        ws.column_dimensions[letter].width = col_widths[i]

    data_font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # --- Hoja Planner ---
    ws2 = wb.create_sheet(SHEET_PLANNER)
    planner_headers = ["Representante", "Área", "Tarea (Oficio)", "Fecha Oficio",
                       "Fecha Plazo", "Estado", "Resumen"]
    ws2.append(planner_headers)
    style_header(ws2, len(planner_headers))

    for o in all_oficios:
        if o.get("representante"):
            ws2.append([
                o.get("representante", ""),
                o.get("area", ""),
                o.get("nombre_oficio", ""),
                format_date_cl(o.get("fecha_oficio")),
                format_date_cl(o.get("fecha_plazo")),
                estado_oficio(o.get("fecha_plazo")),
                o.get("resumen", ""),
            ])

    planner_widths = [26, 24, 28, 14, 14, 18, 55]
    planner_letters = ["A", "B", "C", "D", "E", "F", "G"]
    for i, letter in enumerate(planner_letters):
        ws2.column_dimensions[letter].width = planner_widths[i]

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(planner_headers)):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(EXCEL_FILE)
    return EXCEL_FILE


def confirm_or_correct_area(area_propuesta, confianza):
    """
    Pide al usuario confirmar o corregir el área clasificada.
    Si la confianza es baja, lo señala explícitamente.
    """
    confianza_pct = f"{confianza:.0%}" if confianza else "?"
    low_confidence = confianza and confianza < 0.7

    if low_confidence:
        print(f"    ⚠️  Confianza baja ({confianza_pct}) — revisa con cuidado")

    print(f"    Área propuesta: {area_propuesta} (confianza: {confianza_pct})")
    print(f"    ¿Es correcto? [S/n/cambiar]")
    print(f"      S o Enter = confirmar")
    print(f"      n = ver opciones para cambiar")

    while True:
        resp = input("    → ").strip().lower()

        if resp in ("", "s", "si", "sí", "y", "yes"):
            return area_propuesta

        if resp in ("n", "no", "cambiar", "c"):
            print(f"    Selecciona el área correcta:")
            for i, area in enumerate(AREAS_VALIDAS, 1):
                marker = " ← propuesta" if area == area_propuesta else ""
                print(f"      {i}. {area}{marker}")

            while True:
                try:
                    sel = input(f"    Número (1-{len(AREAS_VALIDAS)}): ").strip()
                    idx = int(sel) - 1
                    if 0 <= idx < len(AREAS_VALIDAS):
                        nueva_area = AREAS_VALIDAS[idx]
                        if nueva_area != area_propuesta:
                            print(f"    ✏️  Corregido: {area_propuesta} → {nueva_area}")
                        return nueva_area
                except (ValueError, KeyboardInterrupt):
                    return area_propuesta
                print(f"    Opción inválida.")

        # Si escribe directamente un número
        try:
            idx = int(resp) - 1
            if 0 <= idx < len(AREAS_VALIDAS):
                nueva_area = AREAS_VALIDAS[idx]
                if nueva_area != area_propuesta:
                    print(f"    ✏️  Corregido: {area_propuesta} → {nueva_area}")
                return nueva_area
        except ValueError:
            pass

        print("    Ingresa S, n, o un número de área.")


def select_representante(area):
    """Permite al usuario seleccionar un representante del área."""
    reps = AREAS_REPRESENTANTES.get(area, [])
    if not reps:
        return ""
    if len(reps) == 1:
        print(f"    Representante asignado: {reps[0]}")
        return reps[0]

    print(f"    Representantes disponibles para {area}:")
    for i, r in enumerate(reps, 1):
        print(f"      {i}. {r}")
    while True:
        try:
            sel = input(f"    Selecciona (1-{len(reps)}, o Enter para omitir): ").strip()
            if not sel:
                return ""
            idx = int(sel) - 1
            if 0 <= idx < len(reps):
                return reps[idx]
        except (ValueError, KeyboardInterrupt):
            return ""
        print(f"    Opción inválida. Ingresa un número entre 1 y {len(reps)}.")


# ============================================================
# MAIN
# ============================================================

def main():
    # Parsear argumentos
    parser = argparse.ArgumentParser(description="Clasificador de Oficios CGE v2")
    parser.add_argument("--stats", action="store_true",
                        help="Mostrar métricas de accuracy del agente")
    parser.add_argument("--regenerar", action="store_true",
                        help="Regenerar reglas de clasificación manualmente")
    args = parser.parse_args()

    # Comando: stats
    if args.stats:
        historial = load_historial()
        if not historial:
            print("\n  No hay historial aún. Procesa algunos oficios primero.\n")
        else:
            print_stats(historial)
        sys.exit(0)

    # Comando: regenerar reglas
    if args.regenerar:
        api_key = get_api_key()
        client = anthropic.Anthropic(api_key=api_key)
        historial = load_historial()
        print("\n  🔄 Regenerando reglas de clasificación...")
        reglas = generate_reglas(client, historial)
        if reglas:
            print("\n  ✅ Reglas regeneradas y guardadas en", REGLAS_FILE)
            print("\n  Reglas generadas:")
            print("  " + "─" * 50)
            for line in reglas.split("\n"):
                print(f"  {line}")
            print("  " + "─" * 50)
        sys.exit(0)

    # Flujo principal
    print()
    print("╔══════════════════════════════════════════════════════════╗")
    print("║     CLASIFICADOR DE OFICIOS CGE — v2 (con aprendizaje)  ║")
    print("╚══════════════════════════════════════════════════════════╝")
    print()

    api_key = get_api_key()
    client = anthropic.Anthropic(api_key=api_key)

    # Cargar historial de aprendizaje
    historial = load_historial()
    if historial:
        stats = compute_stats(historial)
        print(f"  🧠 Historial cargado: {stats['total']} oficios, "
              f"accuracy {stats['accuracy']:.0%}")
        if os.path.exists(REGLAS_FILE):
            print(f"  📏 Reglas aprendidas: cargadas desde {REGLAS_FILE}")
        else:
            print(f"  📏 Reglas aprendidas: se generarán al alcanzar {REGLAS_CADA_N} oficios")
    else:
        print("  🧠 Sin historial previo — el agente empezará a aprender desde cero")
    print()

    # Buscar PDFs
    pdfs = find_oficios_pdfs()
    print(f"  📂 Directorio: {os.getcwd()}")
    print(f"  📄 PDFs encontrados (Ord/OC/RE): {len(pdfs)}")

    if not pdfs:
        print()
        print("  No se encontraron archivos PDF que comiencen con Ord, OC o RE.")
        print("  Asegúrate de que este script esté en la misma carpeta que los oficios.")
        sys.exit(0)

    # Cargar existentes del Excel
    existing = load_existing_oficios()
    new_pdfs = [p for p in pdfs if p not in existing]

    print(f"  ✅ Ya procesados: {len(existing)}")
    print(f"  🆕 Nuevos por procesar: {len(new_pdfs)}")

    if not new_pdfs:
        print()
        print("  No hay oficios nuevos. Todos ya están en el Excel.")
        print("  Tip: Ejecuta con --stats para ver métricas del agente.")
        sys.exit(0)

    print()
    print("  " + "─" * 54)

    # Procesar nuevos
    nuevos_oficios = []
    oficios_procesados_esta_sesion = 0

    for i, pdf in enumerate(new_pdfs, 1):
        print(f"\n  📋 [{i}/{len(new_pdfs)}] Procesando: {pdf}")

        try:
            result = classify_pdf(client, pdf, historial)
            result["archivo"] = pdf

            print(f"    Nombre:    {result.get('nombre_oficio', '?')}")
            print(f"    Resumen:   {result.get('resumen', '?')}")
            print(f"    Keywords:  {', '.join(result.get('keywords', []))}")
            print()

            # === FEEDBACK LOOP ===
            area_propuesta = result.get("area", "Reclamos")
            confianza = result.get("confianza", None)

            area_final = confirm_or_correct_area(area_propuesta, confianza)
            result["area"] = area_final

            # Seleccionar representante
            representante = select_representante(area_final)
            result["representante"] = representante

            # Guardar en historial
            add_to_historial(historial, result, area_propuesta, area_final, representante)

            nuevos_oficios.append(result)
            oficios_procesados_esta_sesion += 1

        except KeyboardInterrupt:
            print("\n\n  ⏹️  Procesamiento interrumpido por el usuario.")
            break
        except json.JSONDecodeError as e:
            print(f"    ❌ Error al parsear respuesta JSON: {e}")
            print(f"    Saltando este oficio...")
            continue
        except Exception as e:
            print(f"    ❌ Error procesando {pdf}: {e}")
            print(f"    Saltando este oficio...")
            continue

    # Verificar si hay que regenerar reglas
    total_procesados = len(historial)
    reglas_exist = os.path.exists(REGLAS_FILE)
    should_regenerate = (
        oficios_procesados_esta_sesion > 0
        and total_procesados >= REGLAS_CADA_N
        and (
            not reglas_exist
            or total_procesados % REGLAS_CADA_N < oficios_procesados_esta_sesion
        )
    )

    if should_regenerate:
        print()
        print(f"  🔄 {total_procesados} oficios procesados — regenerando reglas de clasificación...")
        reglas = generate_reglas(client, historial)
        if reglas:
            print(f"  ✅ Reglas actualizadas en {REGLAS_FILE}")

    # Reconstruir lista completa para Excel
    # Leer oficios existentes del Excel actual si existe
    all_oficios = []
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE, data_only=True)
            if SHEET_OFICIOS in wb.sheetnames:
                ws = wb[SHEET_OFICIOS]
                headers_row = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # tiene archivo
                        o = {}
                        for h_idx, h_name in enumerate(headers_row):
                            if h_idx < len(row):
                                o[h_name] = row[h_idx]
                        # Mapear a formato interno
                        all_oficios.append({
                            "archivo": o.get("Archivo", ""),
                            "nombre_oficio": o.get("Nombre Oficio", ""),
                            "remitente": o.get("Remitente", ""),
                            "fecha_oficio": o.get("Fecha Oficio", ""),
                            "fecha_plazo": o.get("Fecha Plazo", ""),
                            "area": o.get("Área", ""),
                            "representante": o.get("Representante", ""),
                            "resumen": o.get("Resumen", ""),
                        })
            wb.close()
        except Exception:
            pass

    all_oficios.extend(nuevos_oficios)

    # Generar Excel
    if nuevos_oficios:
        output = build_excel(all_oficios)
        print()
        print("  " + "─" * 54)
        print(f"\n  ✅ Excel generado: {output}")
        print(f"     📊 Total oficios: {len(all_oficios)}")
        print(f"     🆕 Nuevos agregados: {len(nuevos_oficios)}")

        # Mostrar resumen de correcciones de esta sesión
        correcciones = sum(
            1 for h in historial[-oficios_procesados_esta_sesion:]
            if h.get("fue_corregido")
        )
        if correcciones:
            print(f"     ✏️  Correcciones realizadas: {correcciones}")

        stats = compute_stats(historial)
        print(f"     🧠 Accuracy acumulada: {stats['accuracy']:.0%} "
              f"({stats['correctos']}/{stats['total']})")
    else:
        print("\n  No se procesaron oficios nuevos en esta sesión.")

    print()
    print("  Tip: Ejecuta con --stats para ver métricas detalladas")
    print("  Tip: Ejecuta con --regenerar para forzar actualización de reglas")
    print()


if __name__ == "__main__":
    main()
